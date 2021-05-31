# -*- coding: utf-8 -*-
"""
Created on Mon Feb 24 20:05:02 2020

@author: Master
"""
import os, re
import shutil
import xlsxwriter
import pandas as pd
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook

def get_save_directory():
    root.filelink = filedialog.askdirectory()    
    link_save_box.config(state='normal')
    link_save_box.delete(0, END)
    global link_save
    link_save = root.filelink.replace('/', '\\\\')
    link_save_box.insert(0, link_save.replace( '\\\\', '\\'))   
    link_save_box.config(state='disabled')

def copy_outline_file_to_save_folder():    
    if os.path.exists(link_save + '\\\\Outline.xlsx'):
        pass
    else:
        link_save_temp = link_save.replace( '\\\\', '\\')
        shutil.copy(os.path.abspath('Outline.xlsx'),link_save_temp + '\\Outline.xlsx')
        
def the_min(lst):
    min_lst = min(lst)
    return lst.index(min_lst)

def file_ST_open():
    root.filename = filedialog.askopenfilename(initialdir = "/",
                                               title = "Select A File",
                                               filetype = (('log', '*.log'),
                                                           ('text', '*.txt'),
                                                           ('All Shit', '*.*')))
    
    number_of_ST.config(state='normal')
    link_ST_box.config(state='normal')
    
    link_ST_box.delete(0, END)
    number_of_ST.delete(0, END)
    
    global ST_num, ST_4, X_max, Y_max, X_min, Y_min
    
    ST_X    = list()
    ST_Y    = list()
    X_max = list()
    X_min = list()
    Y_max = list()
    Y_min = list() 
    
    link_ST = root.filename.replace('/', '\\\\')   
    link_ST_box.insert(0, link_ST.replace( '\\\\', '\\'))
    link_ST_box.config(state='disabled')
    
    ST_count = 0
    try:
        file = open(link_ST,'r',encoding="utf-8")    
    except UnicodeDecodeError:
        file = open(link_ST,'r')     
    except PermissionError:
        messagebox.showinfo("Permission Error","Can not open ST file. \n Try to save your file to another directory")
        
    for line in file:
        if line.find('X=')>-1:
            ST_count+=1
            ST_X.append(float(re.compile('-*\d+\.\d+').findall(line)[0]))
            ST_Y.append(float(re.compile('-*\d+\.\d+').findall(line)[1]))
    file.close()
    number_of_ST.insert(0, int(ST_count/4))
    if ST_count % 4 !=0:
        messagebox.showinfo("Boundary Error"," ST's boundary is not in rectangle shape. \n Check the number of points in each polyline boundary. \n Must be 4 point each one.")
        
    number_of_ST.config(state='disabled')
    ST_num = int(number_of_ST.get())

##    
    for i in range(0,4*ST_num,4):
        X_max.append(max(ST_X[i:i+4]))
        Y_max.append(max(ST_Y[i:i+4]))
        X_min.append(min(ST_X[i:i+4]))
        Y_min.append(min(ST_Y[i:i+4])) 

##Auto sort ST by JSP present with 2 or 4 graph per page
    if var_6.get()=='no':
        pass
    else:
        list_sort = list(zip(X_max,Y_max,X_min,Y_min))
        df = pd.DataFrame(list_sort, columns = ['X_max','Y_max','X_min','Y_min'])
        df = df.sort_values(by=['Y_max','X_max'], ascending=[False, True])
        
        X_max = list(df['X_max'])
        X_min = list(df['X_min'])
        Y_max = list(df['Y_max'])
        Y_min = list(df['Y_min']) 

# If present in 4ST on One paper        
        if len(set(Y_max)) < len(Y_max):
            ST_4 = list()
            for i in range(0,ST_num,4):
                ST_4.append(i)
                if i+2 < ST_num:
                    ST_4.append(i+2)
                if i+1 < ST_num:
                    ST_4.append(i+1)
                if i+3 < ST_num:
                    ST_4.append(i+3)

def file_object_open():
    root.filename = filedialog.askopenfilename(initialdir = "/",
                                               title = "Select A File",
                                               filetype = (('log', '*.log'),
                                                           ('text', '*.txt'),
                                                           ('All', '*.*')))
    
    number_of_Object.config(state='normal')
    link_check_box.config(state='normal')
    link_object_box.config(state='normal')
    
    number_of_Object.delete(0, END)
    link_object_box.delete(0, END)
    link_check_box.delete(0, END)
    
    global link_object
    link_object = root.filename.replace('/', '\\\\')
    link_object_box.insert(0, link_object.replace( '\\\\', '\\'))
    link_object_box.config(state='disabled')   
    Object_count = 0  
    
    try:
        file = open(link_object,'r',encoding="utf-8")
        for line in file:
            if line.find('X=')>-1:
                Object_count+=1
    except UnicodeDecodeError:
        file = open(link_object,'r')
        for line in file:
            if line.find('X=')>-1:
                Object_count+=1
    except PermissionError:
        messagebox.showinfo("Permission Error","Can not open object file. \n Try to save your file to another directory")
    
    number_of_Object.insert(0, Object_count)    
    number_of_Object.config(state='disabled')

def file_check_open():
    root.filename = filedialog.askopenfilename(initialdir = "/",
                                               title = "Select A File",
                                               filetype = (('log', '*.log'),
                                                           ('text', '*.txt'),
                                                           ('All Shit', '*.*')))

    number_of_Object.config(state='normal')    
    link_object_box.config(state='normal')
    link_check_box.config(state='normal')
    
    link_object_box.delete(0, END)
    link_check_box.delete(0, END)
       
    global link_check
    
    link_check = root.filename.replace('/', '\\\\')
    link_check_box.insert(0, link_check.replace( '\\\\', '\\'))
    link_check_box.config(state='disabled')

    Object_count = 0
    try:
        file = open(link_check,'r',encoding="utf-8")
    except UnicodeDecodeError:
        file = open(link_check,'r')
    except PermissionError:
        messagebox.showinfo("Permission Error","Can not open check file. \n Try to save your file to another directory")
      
    for line in file:
        if line.find('Layer:')>-1:
            Object_count+=1
    number_of_Object.delete(0, END)
    number_of_Object.insert(0, Object_count)
    
    number_of_Object.config(state='disabled')

def EPS():
   
    list_of_key = {3:'X=',4:'text',1:'Layer:',2:'Color'}
           
    try:
        file = open(link_object,'r',encoding="utf-8")
        Lines = file.readlines()
    except UnicodeDecodeError: 
        file = open(link_object,'r')
        Lines = file.readlines()
    keyword     = "".join(set(var_2.get()+var_1.get()))     
    file.close()
    
    Coordinate  = list()
    Content     = list()
    Layer       = list()
    Color       = list()
    Color_id    = list()
    items       = [0, Layer, Color, Coordinate, Content]
    count       = [0, 0, 0, 0, 0]       
    X           = list()
    Y           = list()
       
    for line in Lines:
        for key in keyword:
            if line.find(list_of_key[int(key)])>-1:
                count[int(key)]+=1
    #Add information to items
                if     int(key) ==3:
                    X.append(float(re.compile('-*\d+\.\d+').findall(line)[0]))
                    Y.append(float(re.compile('-*\d+\.\d+').findall(line)[1]))                    
                elif   int(key) ==2:
                    if line.find('BYBLOCK')!=-1:
                        pass
                    else:
                        Color_id.append(count[3])                    
                        items[2].append(re.compile('\d+').findall(line)[0])                                    
                else:
                    items[int(key)].append(line.replace(' ','').replace('\n',''))

    #items handling
    #Layer string extracting
    tem_range = len(Layer)
    for i in range(tem_range):
        Layer[i] = Layer[i].split('"')[1]
    
    Length      = list()
    Acreage     = list()
    Wall_A      = list()
    Cut_A       = list()
    Total_A     = list()
    EPS_type    = list()
    
    tokens = ['L=','A=','③', 'U+2462', '②', 'U+2461', '①', 'U+2460', 't=']
    # id stored
    tem_range = len(Content)
    for i in range(tem_range):
        if tokens[0] in Content[i]:
            Content[i] = float(re.compile('-*\d+\.\d+').findall(Content[i])[0])
            Length.append(i)
        elif tokens[1] in Content[i]:
            Content[i] = float(re.compile('-*\d+\.\d+').findall(Content[i])[0])
            Acreage.append(i)
        elif tokens[2] in Content[i] or tokens[3] in Content[i]:
            Content[i] = float(re.compile('-*\d+\.\d+').findall(Content[i])[0])
            Wall_A.append(i)
        elif tokens[4] in Content[i] or tokens[5] in Content[i]:
            try:
                Content[i] = float(re.compile('-*\d+\.\d+').findall(Content[i])[0])
                Cut_A.append(i)
            except Exception:
                Content[i] = Content[i].replace('=.','=0.')
                Content[i] = float(re.compile('-*\d+\.\d+').findall(Content[i])[0])
                Cut_A.append(i)               
                
        elif tokens[6] in Content[i] or tokens[7] in Content[i]:
            Content[i] = float(re.compile('-*\d+\.\d+').findall(Content[i])[0])
            Total_A.append(i)
        elif tokens[8] in Content[i]:
            Content[i] = Content[i].replace('-',' ').replace('t=',' ').replace('H','')
            EPS_type.append(i)        
    
    #ST classify
    fr_de = int(ST_correction_box.get()) #frame deviation
    ST = list()
    
    tem_range = len(X)
    for i in range(tem_range):
        Out_check = list()
        for k in range(ST_num):
            if (X_max[k] + fr_de > X[i] > X_min[k] - fr_de) and (Y_max[k] + fr_de > Y[i] > Y_min[k] - fr_de):
                
    # If present in: 4ST on One paper   
                if len(set(Y_max)) < len(Y_max) and var_6.get() != 'no':
                    ST.append(ST_4[k])
                    Out_check.append(1)
    # If present in: 2ST on One paper
                else:    
                    ST.append(k)
                    Out_check.append(1)
        if len(Out_check) == 0:
            nearest_ST = list()
            for k in range(ST_num):
                X_ce = 0.5*X_max[k] + 0.5*X_min[k]
                Y_ce = 0.5*Y_max[k] + 0.5*Y_min[k]
                nearest_ST.append(abs((X_ce-X[i])**2 + (Y_ce-Y[i])**2))
                if nearest_ST[k] == min(nearest_ST):
                    wrong_ST = k
            Out_Error_Message = " %s is out of ST%i's boundary" %(Content[i],wrong_ST+int(ST_box.get()))
            messagebox.showinfo("Error", Out_Error_Message)
            file.close()

    #Return the true ST
    tem_range = len(ST)
    for i in range(tem_range):
        ST[i]+= int(ST_box.get()) #start_ST: can be replaced
    
    code = list()
    for i in range(count[3]):
        code.append(0)
    
    #EPS_type classify                
    #1_Find the two closest '③' & '①'  text of EPS_type text   
   
    EPS_type_1 = list() #Total_A
    EPS_type_3 = list() #Wall_A
    
    for i in EPS_type:
        Xs = X[i]
        Ys = Y[i]
    
        interval = list() #distance between the EPS_type and ① in the same ST
        k_store = list() #save the id of ① in the same ST
        for k in Total_A:
            if ST[k] == ST[i] and Ys > Y[k] and abs(Xs - X[k])<1000 and abs(Ys - Y[k])<1200:
                interval.append( (Xs-X[k])**2 + (Ys-Y[k])**2 )
                k_store.append(k)
        if len(interval) == 0:
            EPS_type_1.append(-1)
        elif len(interval) != 0:
            EPS_type_1.append(k_store[the_min(interval)])
    
        interval = list() #distance between the EPS_type and ③ in the same ST
        k_store = list() #save the id of ③ in the same ST
        for k in Wall_A:
            if ST[k] == ST[i] and Ys > Y[k] and abs(Xs - X[k])<1000 and abs(Ys - Y[k])<2400:
                interval.append( (Xs-X[k])**2 + (Ys-Y[k])**2 )
                k_store.append(k) 
        if len(interval) == 0:
            EPS_type_3.append(-1)
        elif len(interval) != 0:
            EPS_type_3.append(k_store[the_min(interval)])
         
    #2_Determine the EPS_type thickness by Color (Object: ②, Length, Acreage of Dewatering part)
    
    try:
        file = open(os.path.abspath('Code_dict.txt'),'r',encoding="utf-8")
    except UnicodeDecodeError:
        file = open(os.path.abspath('Code_dict.txt'),'r')
    code_dict={}
    for line in file:
        if line[0] != '#':
            line = line.replace('\n','')
            code_dict[line.split(':')[0]] = line.split(':')[1];
    file.close()              
    
    ST_250 = list()
    for i in EPS_type:
        if "250" in Content[i]:
            ST_250.append(ST[i])
    for i in Color_id:
        code[i] = code_dict[Color[Color_id.index(i)]] + str(ST[i])
    for i in EPS_type_1:
        if i!= -1:
            try:
                code[i] = '1' + Content[EPS_type[EPS_type_1.index(i)]].split(' ')[1] + Content[EPS_type[EPS_type_1.index(i)]].split(' ')[2] + str(ST[i])
            except IndexError:
                pass
    for i in EPS_type_3:
        if i!= -1:
            try:
                code[i] = '3' + Content[EPS_type[EPS_type_3.index(i)]].split(' ')[1] + Content[EPS_type[EPS_type_3.index(i)]].split(' ')[2] + str(ST[i])
            except IndexError:
                pass
    
    for i in list(set(Cut_A) - (set(Color_id))):
        code[i] = '220500' + str(ST[i])
    for i in list(set(Length) - (set(Color_id))):
        code[i] = 'L500' + str(ST[i])
    for i in list(set(Acreage) - (set(Color_id))):
        code[i] = 'A500' + str(ST[i])
    for i in list(set(Length) & (set(Color_id))):
        code[i] = 'L250' + str(ST[i])
    for i in list(set(Acreage) & (set(Color_id))):
        code[i] = 'A250' + str(ST[i])
    
    type = list()
    for i in code:
        type.append(str(i)[0])
    
    wb2 = load_workbook(link_save + '\\\\Outline.xlsx')

    if var_2.get()=='134':
        ws2=wb2['WB_Input']
        ws2.delete_cols(1,15)
        tem_range = len(Content)
        for i in range(tem_range):
            Content[i]=Content[i].replace('text','')
            if Layer[i].upper().find(Content[i].upper())==-1:
                 messagebox.showinfo("Error!!! Mismatched","At: \tST%s\nLayer: \t%s\nBlock: \t%s                    " %(ST[i],Layer[i],Content[i]))
                
    else:
        ws2=wb2['EPS_rawinput']
        ws2.delete_cols(1,15)
        wb2['EPS'].cell(1,5).value = max(ST)
        if len(ST_250) !=0:
            wb2['EPS'].cell(2,5).value = min(ST_250)
            wb2['EPS'].cell(3,5).value = max(ST_250)
        else:
            wb2['EPS'].cell(2,5).value = max(ST)+1
            wb2['EPS'].cell(3,5).value = max(ST)+1
    
    Head = ['Content','ST', 'X', 'Y', 'Code', 'Layer', 'Type']
    tem_range = len(Head)
    for i in range(tem_range):
        ws2.cell(1,i+1).value = Head[i]
    tem_range = len(Content)
    for i in range(tem_range):
        ws2.cell(i+2,1).value = Content[i]
    
    tem_range = len(ST)
    for i in range(tem_range):
        ws2.cell(i+2,2).value = ST[i]
        ws2.cell(i+2,3).value = X[i]
        ws2.cell(i+2,4).value = Y[i]
        ws2.cell(i+2,5).value = code[i]        
        ws2.cell(i+2,7).value = type[i]        
    
    tem_range = len(Layer)
    for i in range(tem_range):
        ws2.cell(i+2,6).value = Layer[i]
        
    wb2.save(link_save + '\\\\Outline.xlsx')
    print('\007')
    
    tem_range = len(Content)
    for  i in range(tem_range): #check_content in Content:
        if str(Content[i]).find('###')>-1:
            messagebox.showinfo("Error","######### is avaiable at ST%s" %(ST[i]))

def check():  

    object={1:'length',2:'Area',3:'X=',4:'text'};
    try:
        a=sum( 1 for line in open(link_check,"r+",encoding="utf-8"))
        file_input = open(link_check,"r+",encoding="utf-8");
    except UnicodeDecodeError:
        a=sum( 1 for line in open(link_check,"r+"))
        file_input = open(link_check,"r+");
          
    X    = list()
    Y    = list()
    XY_line = list()    
    refine_XY = list()
    Y_check = list()
    X_check = list()      
    ST = list()
    Length      = list()
    Area        = list()
    Area_line   = list()
          
    #INPUT
    
    for i in range(0,a):
        data_input = file_input.readline()
        if(data_input.find(object[1])>-1):
            Length.append(float(re.compile('-*\d+\.\d+').findall(data_input)[0]))
        if(data_input.find('Length')>-1):
            Length.append(float(re.compile('-*\d+\.\d+').findall(data_input)[0]))
        if(data_input.find('area')>-1):
            try:
                Area.append(float(re.compile('-*\d+\.\d+').findall(data_input)[0]))
                Area_line.append(i)
            except IndexError:
                Area.append(0)
                Area_line.append(i)
        if data_input.find('X=')>-1 and data_input.find('center') == -1 :
            XY_line.append(i)              
            X.append(float(re.compile('-*\d+\.\d+').findall(data_input)[0]))
            Y.append(float(re.compile('-*\d+\.\d+').findall(data_input)[1]))
    
    #Select 1 XY to determine ST of an area
    tem_range = len(Area_line)
    for i in range(tem_range):        
        for k in XY_line:
            try:
                if Area_line[i] < k < Area_line[i+1] and len(refine_XY) == i:              
                    refine_XY.append(k)
            except IndexError:
                pass
            if k > Area_line[-1] and len(refine_XY) == i:
                refine_XY.append(k)
    
    #ST classify, ST of Area
    fr_de = int(ST_correction_box.get()) #frame deviation
    for i in refine_XY:
        for k in range(ST_num):
            if X_max[k] + fr_de > X[XY_line.index(i)] > X_min[k] - fr_de and Y_max[k] + fr_de > Y[XY_line.index(i)] > Y_min[k] - fr_de:
                if len(set(Y_max)) < len(Y_max) and var_6.get() != 'no':
                    ST.append(ST_4[k]+int(ST_box.get()))
                else:
                    ST.append(k+int(ST_box.get()))
        Y_check.append(Y[XY_line.index(i)])
        X_check.append(Y[XY_line.index(i)])
    
    #OUTPUT            
    Head = ['Length','Area','ST','Sum','','Line_number','X','Y']
    

    wb2 = load_workbook(link_save + '\\\\Outline.xlsx')

    ws2=wb2['PL_Area_CHECK']
    ws2.delete_cols(1,15)
    
    tem_range = len(Head)
    for i in range(tem_range):
        ws2.cell(1,i+1).value = Head[i]
    tem_range = len(Length)
    for i in range(tem_range):
        ws2.cell(i+2,1).value = Length[i]/1e3    
    tem_range = len(Area)
    for i in range(tem_range):
        ws2.cell(i+2,2).value = Area[i]
        ws2.cell(i+2,3).value = ST[i]
        ws2.cell(i+2,6).value = refine_XY[i]
        ws2.cell(i+2,7).value = X_check[i]
        ws2.cell(i+2,8).value = Y_check[i]        
    
    ws2.cell(3,4).value = sum(Area)/1e6
    ws2.cell(2,4).value = sum(Length)/1e3
    ws2.cell(2,5).value = 'Length'
    ws2.cell(3,5).value = 'Area'
        
    wb2.save(link_save + '\\\\Outline.xlsx')    
    file_input.close()
    print('\007')

def LA():
    
    wb = xlsxwriter.Workbook(link_save + '\\\\LA.xlsx')
    sheet1 = wb.add_worksheet(var_3.get() + var_4.get())
    
    # object={1:'length',2:'area',3:'perimeter'};
    Length      = list()
    Area        = list()
    if var_3.get() == 'L':
        options = '1'
    if var_4.get() == 'A':
        options = '2'
    try:
        file = open(link_object,"r+",encoding="utf-8");
        for line in file:
            if options == '2' and line.find('area')>-1:
                Area.append(float(re.compile('-*\d+\.\d+').findall(line)[0])/1e6)
            if options == '1' and (line.find('Length')>-1 or line.find('length')>-1 or line.find('perimeter')>-1):
                Length.append(float(re.compile('-*\d+\.\d+').findall(line)[0]))
    except UnicodeDecodeError:
        file = open(link_object,"r+")
        for line in file:
            if options == '2' and line.find('area')>-1:
                Area.append(float(re.compile('-*\d+\.\d+').findall(line)[0])/1e6)
            if options == '1' and (line.find('Length')>-1 or line.find('length')>-1 or line.find('perimeter')>-1):
                Length.append(float(re.compile('-*\d+\.\d+').findall(line)[0]))

               
    Head = ['Length','Area', 'Sum']
    sheet1.write_row(0,0,Head)
    sheet1.write_column(1,0,Length)
    sheet1.write(1,2,sum(Length)/1e3)
    sheet1.write(2,2,sum(Area))
    sheet1.write_column(1,1,Area)        
    file.close()
    wb.close()

def process():
    
    copy_outline_file_to_save_folder()
    if var_5.get()=='check':
        check()
        check_but_1.deselect()
        check_but_2.deselect()    
        check_but_3.deselect()
        check_but_4.deselect()

        os.startfile(link_save + '\\\\Outline.xlsx')

    if var_3.get()=='L' and var_4.get()=='A':
        check_but_3.deselect()
        check_but_4.deselect()
    if var_1.get()=='234' or var_2.get()=='134':
        EPS()
        os.startfile(link_save + '\\\\Outline.xlsx')
        
    if var_3.get()=='L' or var_4.get()=='A':
        LA()
        os.startfile(link_save + '\\\\LA.xlsx')

###### GUI design

root = Tk()
root.resizable(width = 0, height = 0)
root.title("Copyright 2021 © Eng.Nhat Quang")
root.minsize(450, 350)
root.iconbitmap(os.path.abspath('iconF.ico').replace('\\','\\\\'))
run_img = Image.open(os.path.abspath('play.png').replace('\\','\\\\'))
run_img = run_img.resize((40,40), Image.ANTIALIAS)
photoImg =  ImageTk.PhotoImage(run_img) 

# #CHECK BUTTON

var_1 = StringVar()
check_but_1 = Checkbutton(root, text = "Texts",
variable = var_1,
onvalue="234",
offvalue="")
check_but_1 .grid(row = 9, column = 0, padx = 10, sticky = "WN")

var_2 = StringVar()
check_but_2 = Checkbutton(root, text = "Wall Blocks",
variable = var_2,
onvalue="134",
offvalue="")
check_but_2 .grid(row = 10, column = 0, padx = 10, sticky = "WN")

var_3 = StringVar()
check_but_3 = Checkbutton(root, text = "Sum of Length",
variable = var_3,
onvalue="L",
offvalue="")
check_but_3 .grid(row = 9, column = 1, padx = 5, sticky = "WN", columnspan = 2)

var_4 = StringVar()
check_but_4 = Checkbutton(root, text = "Sum of Area",
variable = var_4,
onvalue="A",
offvalue="")
check_but_4 .grid(row = 10, column = 1, padx = 5, sticky = "WN", columnspan = 2)

var_5 = StringVar()
check_but_5 = Checkbutton(root, text = "Check",
variable = var_5,
onvalue="check",
offvalue="")
check_but_5.grid(row = 9, column = 3, padx = 20, sticky = "W", columnspan = 1)

var_6 = StringVar()
check_but_6 = Checkbutton(root, text = "Auto numbering ST",
variable = var_6,
onvalue="yes",
offvalue="no")
check_but_6 .grid(row = 4, column = 0, padx = 10, sticky = "WN")


label_1 = Label(root, text = "ST's border\ncorrection\n(mm):")
label_2 = Label(root, text = "Objects to scan:", font = ('TkDefaultFont', 10, 'bold'))
label_3 = Label(root, text = "Number of\nST:")
label_4 = Label(root, text = "Number of\nObject:")
label_5 = Label(root, text = "First ST's\nnumber")

label_1.grid(row = 19, column = 3, padx = 10, sticky = "W")
label_2.grid(row = 6, column = 0, padx = 10, pady = 12, sticky = "WS")
label_3.grid(row = 19, column = 1, padx = 0, sticky = "W")
label_4.grid(row = 19, column = 0, padx = 10, ipady = 2, sticky = "W")
label_5.grid(row = 19, column = 5, padx = 10, sticky = "W")


#ENTRY_BOX

link_ST_box = Entry(root, width=60, relief=FLAT, highlightbackground="black", highlightthickness=1)
link_ST_box.grid(row = 0, column = 1, padx = 5, ipady = 2, sticky = "W", columnspan = 6)

link_object_box = Entry(root, width=60, relief=FLAT, highlightbackground="black", highlightthickness=1)
link_object_box.grid(row = 1, column = 1, padx = 5, ipady = 2, sticky = "W", columnspan = 6)

link_check_box = Entry(root, width=60, relief=FLAT, highlightbackground="black", highlightthickness=1)
link_check_box.grid(row = 2, column = 1, padx = 5, ipady = 2, sticky = "W", columnspan = 6)

link_save_box = Entry(root, width=60, relief=FLAT, highlightbackground="black", highlightthickness=1)
link_save_box.grid(row = 3, column = 1, padx = 5, ipady = 2, sticky = "W", columnspan = 6)

number_of_ST = Entry(root, width=4, relief=FLAT, highlightbackground="black", highlightthickness=1)
number_of_ST.grid(row = 19, column = 2, sticky = "E")

number_of_Object = Entry(root, width=4, relief=FLAT, highlightbackground="black", highlightthickness=1)
number_of_Object.grid(row = 19, column = 0, padx = 10, sticky = "E")

ST_correction_box = Entry(root, width=5, relief=FLAT, highlightbackground="black", highlightthickness=1)
ST_correction_box.grid(row = 19, column = 4, padx = 5, ipady = 2, sticky = "E")
ST_correction_box.insert(0,0)

ST_box = Entry(root, width=4, relief=FLAT, highlightbackground="black", highlightthickness=1)
ST_box.grid(row = 19, column = 6, padx = 5, ipady = 2, sticky = "E")
ST_box.insert(0,1)

#BUTTON

button_1 = Button(root, text = 'Open ST File',
command=file_ST_open,
background = 'SteelBlue1', borderwidth=4).grid(row = 0, 
column = 0, padx = 10, pady = 2, ipady = 2,
sticky = "WE")

button_2 = Button(root, text = 'Open Object File',
command=file_object_open,
background = 'SteelBlue1', borderwidth=4).grid(row = 1,
column = 0, padx = 10, pady = 2, ipady = 2,
sticky = "WE")

button_3 = Button(root, text = 'Open Check File',
command=file_check_open,
background = 'SteelBlue1', borderwidth=4).grid(row = 2,
column = 0, padx = 10, pady = 2, ipady = 2,
sticky = "WE")

button_4 = Button(root, text = 'Save at',
command=get_save_directory,
background = 'SteelBlue1', borderwidth=4).grid(row = 3,
column = 0, padx = 10, pady = 2, ipady = 2,
sticky = "WE")

button_19 = Button(root, command=process, image = photoImg,
background = 'IndianRed3', highlightbackground="black",
highlightthickness=1, borderwidth=1).grid(row = 20,
column = 0, ipadx = 5, columnspan = 8)

root.mainloop() 